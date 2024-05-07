import cv2
import numpy as np
import tkinter as tk
import tensorflow as tf
from tkinter import ttk
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
from datetime import datetime

class Browser:
    def __init__(self, root):
        self.attendance_list = []
        self.root = root

        self.background_image = Image.open("background.jpg")
        self.background_photo = ImageTk.PhotoImage(self.background_image)

        # Create a label with the background image
        self.background_label = tk.Label(root, image=self.background_photo)
        self.background_label.place(x=0, y=0, relwidth=1, relheight=1)

        self.canvas = tk.Canvas(root, width=640, height=480)
        self.canvas.place(x=60, y=170)

        # Load Haar cascade for face detection
        self.face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

        # Load pre-trained model
        self.save_model = tf.keras.models.load_model("khuonmat.h5")

        self.label_text = ""


    def createWidgets(self):

        self.controlFrame = tk.Frame(self.root)
        self.controlFrame.place(x=810, y=50)

        self.label1 = tk.Label(self.controlFrame, text="Face recognization", font=("Calibri", 25) )
        self.label1.pack()

        self.attendance_button = tk.Button(self.controlFrame, text="Capture Attendance", command=self.capture_attendance)
        self.attendance_button.pack(side=tk.RIGHT)  # Place the button on the right side

        self.attended_label = tk.Label(self.controlFrame, text="", font=("Calibri", 16))
        self.attended_label.pack(side=tk.BOTTOM)  # Place the label at the bottom

    def capture_attendance(self):
        # self.attendance_list.append(self.label_text.get())
        # self.status_label.config(text="Attendance Captured for: " + self.label_text.get())
        # Ghi nhận việc điểm danh và cập nhật nhãn
        self.attendance_list.append("Attended")
        self.attended_label.config(text="Attended")

        # Ghi vào file Excel
        self.write_to_excel(self.label_text)

    def write_to_excel(self, label_text):
        try:
            # Thử mở workbook hiện có nếu tồn tại
            wb = load_workbook('attendance.xlsx')
            sheet = wb.active
        except FileNotFoundError:
            # Nếu không tồn tại, tạo một workbook mới
            wb = Workbook()
            sheet = wb.active
            sheet['A1'] = 'Name'
            sheet['B1'] = 'Date'

        # Thêm dữ liệu mới vào sheet
        sheet.append([label_text, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        # Lưu workbook
        wb.save('attendance.xlsx')

    def update_video_feed(self):
        _, frame = self.webcam.read()
        if frame is not None:
            img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = self.face_cascade.detectMultiScale(gray, 1.3, 5)

            for (x, y, w, h) in faces:
                cv2.rectangle(img, (x, y), (x+w, y+h), (0, 255, 0), 2)
                face_gray = gray[y:y+h, x:x+w]
                face_gray = cv2.resize(face_gray, (100, 100))
                face_gray = np.array(face_gray).reshape((1, 100, 100, 1)) / 255.0
                result = self.save_model.predict(face_gray)
                final = np.argmax(result)
                labels = ["Hao", "Hoang"]
                if result[0][final] > 0.9:
                    self.label_text = labels[final]
                else:
                    self.label_text = ""
                cv2.putText(img, self.label_text, (x+10, y+h+30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)

                # key = cv2.waitKey(1) & 0xFF
                # if key == ord('q'):
                #     self.attendance_list.append(label_text)
                #     self.write_to_excel(label_text)
                    # cv2.putText(img, "Attended", (920, 250), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 3)

            photo = ImageTk.PhotoImage(image=Image.fromarray(img))
            self.canvas.create_image(0, 0, anchor=tk.NW, image=photo)
            self.canvas.photo = photo

        self.root.after(10, self.update_video_feed)

    def main(self):
        self.createWidgets()

        self.webcam = cv2.VideoCapture(0)
        self.webcam.set(3, 640)
        self.webcam.set(4, 480)

        self.update_video_feed()

        self.root.mainloop()

        self.webcam.release()

if __name__ == '__main__':

    root = tk.Tk()
    root.geometry("1280x720")
    root.config(background="white")
    root.resizable(0, 0)
    root.title("Face Recognition Attendance System")
    root.resizable(False, False)

    browser = Browser(root)
    browser.main()
